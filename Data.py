
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
import numpy as np
from matplotlib import pyplot
import seaborn as sns
from matplotlib import cm
import matplotlib.patches as mpatches
import math
import pdfplumber
import re
import os
import shutil

def getdetails(crs,year,part,clg):
    #Change The src variables strings according to your storage
    src=""+crs+str(year)+str(part)+".xlsx"
    xls=pd.ExcelFile(src, engine='openpyxl')
    if clg=='':
        return xls.sheet_names
    elif clg in xls.sheet_names:
        return xls.parse(clg)[['Roll No','Name']].dropna().apply(lambda row: ' - '.join(row.values.astype(str)), axis=1).squeeze().tolist()
    
    else:
        return ["invalid college name"]

def studentreport(Df,rollno,sem=0):
    plt.close()
    DF = Df.loc[rollno,]
    res = dict()
    if sem==0:
        res['rollno']=rollno
        res['name']=DF.iloc[0,0]
        res["perc"]=DF['SGPA'].mean()*10
        res["plot"]=DF[DF.columns[DF.columns.str.startswith("GP")]].plot.barh(stacked=True)
        plt.close()
        plt.pie([100-res["perc"],res["perc"]], labels=[" ","Agg. Percentage"], colors=['#FFB946','#2ED47A'], autopct='%1.1f%%', shadow=False, startangle=190)
        centre_circle = plt.Circle((0,0),0.80,color=None, fc='#F6F8FB',linewidth=1.25)
        fig = plt.gcf()
        fig.gca().add_artist(centre_circle)
        res["percplot"]=fig
        if DF['Result'].notnull().any():
            res["ER"]=','.join(DF[DF['Result'].fillna("0").str.startswith('ER')]['Result'].str.cat().split('ER -'))
        res['Fail']=False
    else :
        a4_dims = (9.5, 4)
        fig, ax = pyplot.subplots(figsize=a4_dims)
        DF=DF.loc[sem]
        res["perc"]=DF['SGPA']*10
        Ser=pd.Series(DF[DF.index[DF.index.str.startswith("GP")]].squeeze().tolist(),index=DF[DF.index[DF.index.str.startswith("Sub")]].squeeze().tolist())
        res["plot"]=sns.barplot(ax=ax, y=Ser.index, x=Ser.values, palette="rocket_r", orient="h")
        res["plot"].set_xlim(0,10)
        res["plot"].set(xlabel='Grade Points', ylabel='Subject Code')
    return res
 
def get_studentReport(crs,year,part,clg,rollno):
   Df=pd.ExcelFile(""+crs+str(year)+str(part)+".xlsx" ,engine='openpyxl' ).parse(clg,index_col=[0,1])
   file_name=""+crs+str(year)+str(part)+".xlsx"
   path=os.path.abspath(file_name)
   #Df['Roll No']=Df['Roll No'].replace({" ":np.nan, "":np.nan}).fillna(method='ffill')
   Secondly=[]
   try:
       Secondly.append(studentreport(Df,rollno))
       for i in range(1,7):
           Secondly.append(studentreport(Df,rollno,i))
   except:
       Secondly=[]
       Secondly.append({"Fail":True})
       Secondly.append({"path":path})
   print(Secondly)
   return Secondly

def get_subw(Df,year):
    plt.close()
    print(Df)
    DF=Df.xs(year*2-1,level=1).append(Df.xs(year*2,level=1))
    SubW=pd.DataFrame()
    for i,j in zip(DF.columns[DF.columns.str.startswith("Sub")],DF.columns[DF.columns.str.startswith("GP")]):
         try:
             SubW=pd.concat([SubW,DF.groupby(i)[j].mean()])
         except:
             return {"DataError":['Quite Possible that one the GP columns have wrong data in this particular sheet. Please proceed to changes accordingly.']}
    SubW=SubW.reset_index()
    SubW=SubW.groupby('index').mean().squeeze().sort_values()
    SubW.dropna(inplace=True)
    fig, ax = plt.subplots(figsize=(10,8), dpi= 33)
    ax.vlines(x=SubW.index.astype(str), ymin=0, ymax=SubW.values, color='firebrick', alpha=0.7, linewidth=4)
    ax.scatter(x=SubW.index.astype(str), y=SubW.values, s=300, color='purple')
    # Title, Label, Ticks and Ylim
    ax.set_title('Subject Wise Aggregate '+'- Year '+str(year), fontdict={'size':22})
    ax.set_ylabel('Mean SGPA')
    ax.set_xticks(SubW.index.astype(str))
    ax.set_xticklabels(SubW.index.astype(str).str.upper(), rotation=60, fontdict={'horizontalalignment': 'right', 'size':12})
    ax.set_ylim(0, 10)
    return ax
 
 
def getcourse(crs,year,part,clg):
    src=""+crs+str(year)+str(part)+".xlsx"
    xls=pd.ExcelFile(src, engine="openpyxl")
    if clg in xls.sheet_names:
        Df=xls.parse(clg,index_col=[0,1])
    else:
        return ["invalid college name"]
    res=dict()
    Df.replace(0,np.nan,regex=True)
    for _,value in Df[Df.columns[Df.columns.str.startswith("Sub")]].iteritems():
        Df[value.name]=Df[value.name].astype(str).str.slice(0,8)
    res['Sub_W']=dict()
    for i in range(1,part+1):
         subw=get_subw(Df,i)
         if type(subw)==dict:
              print(type(subw))
              return subw
         else:
              res['Sub_W'][i]=subw
    SubW=pd.DataFrame([],columns=['Roll No','Name','FirstYear','SecondYear','ThirdYear','Total'])
    for _, value in Df.groupby(level="Roll No"):
        DF=value['SGPA'].squeeze()
        SubW.loc[len(SubW.index)]=[str(value.index.get_level_values(0).unique().values[0]),value.iloc[0]["Name"],round(DF[0:2].mean(),2),round(DF[2:4].mean(),2),round(DF[4:6].mean(),2),round(DF[0:6].mean(),2)]
    res['Stu_W']=SubW.sort_values(by='Total',ascending=False)
    plt.close()
    plt.figure(figsize=(5.2,3.6), dpi=60)
    ER_roll=Df[Df['Result'].notnull()].index.get_level_values(0).unique()
    res['ER']=ER_roll
    Df=Df.drop(ER_roll)
    P1=Df.groupby('SEM')['SGPA'].mean().plot.barh(color='skyblue')
    P1.set_xlabel("MEAN SGPA")
    P1.set_xlim(0,10)
    P1.set_title('Semester Wise Aggregate', fontdict={'size':12})
    res['Sem_W']=P1
    plt.close()
    res["perc"]=Df['SGPA'].mean()*10
    plt.figure(figsize=(4,4), dpi=60)
    plt.pie([100-res["perc"],res["perc"]], labels=[" ","Agg. Percentage"], colors=['#F3F3F3','#b300b3'],autopct='%1.1f%%', shadow=True)
    centre_circle = plt.Circle((0,0),0.75,color='white', fc='white',linewidth=1.25)
    fig = plt.gcf()
    fig.gca().add_artist(centre_circle)
    res["percplot"]=fig
    plt.close()
    return res

def get_rank(crs,year,part=3,clg=None,campus='All'):
     src=''+crs+str(year)+str(part)+"rank.xlsx"
     Df=pd.read_excel(src,sheet_name=campus,index_col=[0,1])
     res=dict()
     plt.close()
     if clg :
          Df=Df.filter(like=clg,axis=0)
          plt.close()
          fig=plt.figure(figsize=(6,5), dpi=60)
          res["perc"]=len(Df.index)
          plt.pie([100-res["perc"],res["perc"]], colors=['#F3F3F3','#5FE48D'],autopct='%.0f', shadow=False,
                  startangle=110, textprops={'fontsize':12}, explode=[0, 0.1],
                  labels=['Others',''], labeldistance=0.4)
          #centre_circle = plt.Circle((5,5),1,color='white', fc='white',linewidth=1.25)
          fig = plt.gcf()
          #fig.gca().add_artist(centre_circle)
          res["plot"]=fig
          plt.close()
     else:
          fig=plt.figure(figsize=(6,7), dpi=60)
          data=Df.groupby(level=[0]).size()
          res["plot"]=data.plot.bar(color="#45B39D")
          plt.tight_layout()
          red_patch = mpatches.Patch(color='#45B39D', label='No. of Students in top 100')
          plt.legend(handles=[red_patch])
          plt.yticks([i for i in range(0,31,2)])
          plt.xticks(rotation=90)
          plt.close()
          
     plt.close()
     res["df"]=Df
     return res
def create_rank(crs,year,sem):
     src=""+crs+str(year)+str(sem)+".xlsx"
     Df=pd.concat(pd.read_excel(src,sheet_name=None, engine="openpyxl"),axis=0)
     Df["GR.CGPA"]=np.nan
     for _,i in Df[Df['Name'].notnull()].iterrows():
          hee=Df.index.get_loc(i.name)
          
          hehe=Df.columns.get_loc("GR.CGPA")
          DF=Df.iloc[hee:hee+(sem*2)]
          #print(DF["Roll No"].astype(str))
          Df.iloc[hee,hehe]=DF['CGPA'].mean()
     Df=Df[Df.Name.notnull()][['Roll No','Name','GR.CGPA']].sort_values("GR.CGPA",ascending=False)
     North=('Bhagini Nivedita College',
            'Hans Raj College',
            'Hindu College',
            'Kalindi College',
            'Keshav Mahavidyalaya',
            'Kirori Mal College',
            'Maharaja Agrasen College',
            'Miranda House ',
            'Rajdhani College',
            'Ramjas College ',
            'S.G.T.B. Khalsa College',
            'Shivaji College ',
            'Shyam Lal College (Day)',
            'St. Stephens College',
            'Swami Shraddhanand College',
            'Zakir Husain Delhi College (Day',
            'I.P.College For Women',
            'Keshav Mahavidyalaya',
            'Shyama Prasad Mukherjee College',
            'Sri Guru Gobind Singh College',)
     
     South=('Acharya Narendra Dev College',
            'Atma Ram Sanatan Dharam College',
            'Deen Dayal Upadhyaya College',
            'Deshbandhu College (Day) ',
            'Dyal Singh College (Day)',
            'Gargi College',
            'Maitreyi College   ',
            'Moti Lal Nehru College (Day) ',
            'Sri Aurobindo College (Day)',
            'Bhaskaracharya College of Appli',
            'College Of Vocational Studies',
            'Ramanujan College',
            'P.G.D.A.V. College (Day)',
            'Ram Lal Anand College (Day)',
            'Aryabhatta College Formerly Ram',
            'Shaheed Rajguru College of Appl',
            'Shaheed Sukhdev College of Busi')
     print(Df)                          
     writer=pd.ExcelWriter(""+crs+str(year)+str(sem)+"rank.xlsx")
     for i in ("All","South","North"):
         if i == "South":
             DF=Df[np.in1d(Df.index.get_level_values(0),South)]
         elif i=="North":
             DF=Df[np.in1d(Df.index.get_level_values(0),North)]
         else :
             DF=Df 
         DF=DF.iloc[:100]
         DF=DF.assign(Rank=range(1,101))
         DF.to_excel(writer,sheet_name=i)
     writer.save()

ts={
    'vertical_strategy': 'text',
    "horizontal_strategy": 'text',
    "explicit_vertical_lines": [],
    "explicit_horizontal_lines": [],
    "snap_tolerance": 2,
    "join_tolerance": 1,
    "edge_min_length": 3,
    "min_words_vertical": 3,
    "min_words_horizontal": 0,
    "keep_blank_chars": False,
    "text_tolerance": 2,
    "text_x_tolerance": 2,
    "text_y_tolerance": 2,
    "intersection_tolerance": 3,
    "intersection_x_tolerance": None,
    "intersection_y_tolerance": None,
}

def pdfupload(crs,year,part,path):
    res=dict()
    roll=''
    sem=part*2
    cur_sem=1
    file2=''+crs+str(year)+str(part)+'.xlsx'
    wb = Workbook()
    pdf=pdfplumber.open(path)
    College=[]
    print(pdf.pages[1].extract_tables(table_settings=ts)[0][0])
    sub_col_no=4
    for page in pdf.pages:
        lpage=page.extract_text().splitlines()
        # Get all the text information of the current page, including the text in the table
        for line in range(0,len(lpage)):
             if lpage[line].__contains__('CAMPUS'):
                 if lpage[line].__contains__(':'):
                     College.append(lpage[line].split(':')[1].split("NORTH")[0].split("SOUTH")[0].strip())
                 else:
                     College.append(lpage[line-1].split(':')[1].split("NORTH")[0].split("SOUTH")[0].strip())
                 del line
                 break
        
        if (College[len(College)-1]!=College[len(College)-2])|(len(College)==1):
             if len(College)!=1:
                 for i in range(2,len(ws['A']),sem):
                     ws.merge_cells('A'+str(i)+':A'+str(i-1+sem))
                 
             ws=wb.create_sheet(College[len(College)-1])
             ls=pdf.pages[1].extract_tables()[0][0]
             ls.insert(2,'SEM')
             ls=list(map(lambda x: x.split('\n')[0], ls))
             ls[1]=ls[1].split(" ")[0]
             ws.append(ls)
        # Get all the text information of the current page, including the text in the table
        # print(page.extract_text())
        
        for table in page.extract_tables(table_settings=ts):
             # print(table)
             for row in table:
                 #print(sub_col_no)
                 rowlist=str(row).replace("[","",).replace("]","").replace("'","").replace("\\n","").split(",")
                 #print(rowlist)
                 if ' Credit' in rowlist:
                     #print('help',rowlist.index(' Credit'))
                     sub_col_no=rowlist.index(' Credit')  
                   
                 elif (" Sub" not in rowlist)&(rowlist[sub_col_no]!=' '):
                     try:
                          if (int(rowlist[sub_col_no])>=11):
                               if sub_col_no==4:
                                   try:
                                       if len(rowlist[0])<5:
                                           rowlist[0]+=rowlist[1]
                                           rowlist[0]=rowlist[0].replace(" ","")
                                           rowlist[1]=rowlist[2]  
                                   except:
                                       print(rowlist)
                                   del rowlist[3]
                               if cur_sem<sem:
                                   rowlist[2]=cur_sem
                                   cur_sem+=1
                               elif cur_sem==sem:
                                   rowlist[2]=cur_sem
                                   cur_sem=1
                               else:
                                   cur_sem=1
                               if rowlist[0]==roll:
                                   pass
                               elif rowlist[0]=="":
                                   rowlist[0]=roll
                               else:
                                   roll=rowlist[0]
                               ws.append(rowlist)
                               #print(rowlist)
                     except:
                          if  re.sub('[^A-Za-z0-9]+', '',rowlist[sub_col_no]).isalpha():
                               pass
                          else:
                               if sub_col_no==4:
                                   try:
                                       if len(rowlist[0])<5:
                                           rowlist[0]+=rowlist[1]
                                           rowlist[0]=rowlist[0].replace(" ","")
                                           rowlist[1]=rowlist[2] 
                                   except:
                                       print()
                                   del rowlist[3]
                               if len(rowlist[3].lstrip().split())==2:
                                   rowlist.insert(4,rowlist[3].lstrip().split()[1])
                               else:
                                   rowlist.insert(4,'')
                               rowlist[3]= rowlist[3].lstrip().split()[0]
                               if cur_sem<sem:
                                   rowlist[2]=cur_sem
                                   cur_sem+=1
                               elif cur_sem==sem:
                                   rowlist[2]=cur_sem
                                   cur_sem=1
                               else:
                                   cur_sem=1
                               if rowlist[0]==roll:
                                   pass
                               elif rowlist[0]==" ":
                                   rowlist[0]=roll
                               else:
                                   roll=rowlist[0]
                               ws.append(rowlist)
                               #print(rowlist)
        #print('---------- split line ----------')
    pdf.close()
    # Save Excel sheet
    del wb['Sheet']
    for i in range(2,len(ws['A']),sem):
       ws.merge_cells('A'+str(i)+':A'+str(i-1+sem))
                 
    wb.save(file2)
    print('Written to excel successfully')
 
    res=columns(file2)
    coll=list() 
    college_file1=open("Files/Colleges.txt","r+")
    lines=college_file1.readlines()
    coll=[i.strip() for i in lines]
    for i in getdetails(crs, year, part, ""):
        if i.strip() not in coll:
            coll.append(i.strip())
    colls=[i+"\n" for i in coll]
    colls[-1]=colls[-1].strip()
    college_file1.close()
    colleg=open("Files/Colleges.txt","w")
    for i in colls:
        colleg.write(i)
    colleg.close()
    
    
    course=list()
    course_file=open("Files/Courses.txt", "r+")
    line=course_file.readlines()
    course=[i.strip() for i in line]
    if crs not in course: 
        course.append(crs)
    courses=[i+"\n" for i in course]
    courses[-1]=courses[-1].strip()
    course_file.close()
    cou=open("Files/Courses.txt", "w")
    for i in courses:
        cou.write(i)
    cou.close()
    try:
        create_rank(crs,year,part)
    except TypeError:
        res['TypeError']=['Some of the columns might have merged together.','The SGPA row might have wrong data.','Please ensure the credibility of the data before further use.']
        
    if not bool(res):
        res['result']=["Successfully uploaded files"]
    return res

def columns(path):
    dict1=dict()
    shutil.copy(path, "i"+path)
    xls=pd.ExcelFile("i"+path, engine="openpyxl")
    xls1=pd.ExcelWriter(path, engine="openpyxl")
    for i in xls.sheet_names:
            df=pd.read_excel(xls, sheet_name=i, engine="openpyxl")
            b=df.columns
            #print("the sheet "+i+"there is :")
            b=list(b)
            b[1],b[2]=b[2],b[1]
            df=df[b]
            ls=["Roll No", "SEM", "Name", "Sub", "GR", "GP", "CRP", "Sub.1", "GR.1", "GP.1", "CRP.1", "Sub.2", "GR.2", "GP.2", "CRP.2", "Sub.3", "GR.3", "GP.3", "CRP.3", "Sub.4", "GR.4", "GP.4", "CRP.4", "TOT CR", "TOT CRP", "SGPA", "CGPA", "Result", "GR.CGPA", "DIV"]
            dict1[i]=check_column_name(ls, b)
            df=df.reindex(df["Roll No", "SEM"])
            df.to_excel(xls1, sheet_name=i, engine="openpyxl")
    xls1.save()
    return dict1
    
def check_column_name(a,b):
    list1=list()
    for i in a:
        if i not in b:
            #print("Column "+i+" is missing from the sheet.")
            list1.append(i)
    return list1
